"""
Document text extraction — PDF, Excel, Word.
"""

import io
import logging
from typing import Optional

logger = logging.getLogger(__name__)


def extract_pdf_text(content: bytes, max_chars: int = 5000) -> Optional[str]:
    """Extract text from PDF bytes using pdfplumber."""
    try:
        import pdfplumber

        with pdfplumber.open(io.BytesIO(content)) as pdf:
            texts = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    texts.append(text)
                if len("\n".join(texts)) > max_chars:
                    break
            return "\n".join(texts)[:max_chars] if texts else None
    except Exception as e:
        logger.error(f"PDF extraction failed: {e}")
        return None


def extract_excel_text(content: bytes, max_chars: int = 5000) -> Optional[str]:
    """Extract text from Excel bytes using openpyxl."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        texts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text:
                    texts.append(row_text)
                if len("\n".join(texts)) > max_chars:
                    break
        wb.close()
        return "\n".join(texts)[:max_chars] if texts else None
    except Exception as e:
        logger.error(f"Excel extraction failed: {e}")
        return None


def extract_docx_text(content: bytes, max_chars: int = 5000) -> Optional[str]:
    """Extract text from Word document bytes using python-docx."""
    try:
        from docx import Document

        doc = Document(io.BytesIO(content))
        texts = []
        for para in doc.paragraphs:
            if para.text.strip():
                texts.append(para.text)
            if len("\n".join(texts)) > max_chars:
                break
        return "\n".join(texts)[:max_chars] if texts else None
    except Exception as e:
        logger.error(f"DOCX extraction failed: {e}")
        return None


def extract_text_from_attachment(
    filename: str,
    content: bytes,
    max_chars: int = 20000,
) -> Optional[str]:
    """Dispatch by filename extension. Returns None if the type isn't supported."""
    name = (filename or "").lower()
    if name.endswith(".pdf"):
        return extract_pdf_text(content, max_chars=max_chars)
    if name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xls"):
        return extract_excel_text(content, max_chars=max_chars)
    if name.endswith(".docx") or name.endswith(".doc"):
        return extract_docx_text(content, max_chars=max_chars)
    # Plain text and markdown — best effort decode.
    if name.endswith(".txt") or name.endswith(".md"):
        try:
            return content.decode("utf-8", errors="replace")[:max_chars]
        except Exception:
            return None
    return None
